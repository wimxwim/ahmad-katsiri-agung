#!/usr/bin/env python3
"""
Multi-Query Scraper — KULINER di Jakarta Selatan
Auto-connect ke Chrome yang udah running
"""
import re, time, subprocess, sys, json
from fuzzywuzzy import fuzz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

SESSION = "scraper"
CHROME_PORT = 9222

def check_chrome_ready():
    """Check if Chrome is running on port 9222"""
    try:
        r = subprocess.run(
            ['curl', '-s', f'http://localhost:{CHROME_PORT}/json/version'],
            capture_output=True, text=True, timeout=3
        )
        if r.returncode == 0 and r.stdout.strip():
            data = json.loads(r.stdout)
            print(f"✅ Chrome ready: {data.get('Browser', 'Unknown')}", flush=True)
            return True
    except:
        pass
    return False

def find_direct_browser():
    """Find the chrome-direct browser ID from browser-act list"""
    r = subprocess.run(["browser-act", "browser", "list"],
                       capture_output=True, text=True, timeout=5)
    if r.returncode != 0:
        return None
    for line in r.stdout.split('\n'):
        if 'chrome-direct' in line and 'id=' in line:
            m = re.search(r'id=(\S+)', line)
            if m:
                return m.group(1)
    return None

def init_browser():
    """Connect ke Chrome yang sudah running"""
    print("▶ Connecting ke Chrome...", flush=True)

    # Close old session
    subprocess.run(["browser-act", "session", "close", SESSION],
                   capture_output=True, timeout=5)
    time.sleep(1)

    # Auto-discover chrome-direct browser ID
    browser_id = find_direct_browser()
    if not browser_id:
        print("❌ No chrome-direct browser found in browser-act", flush=True)
        print("   Register dulu: browser-act browser create --type chrome-direct --name 'my-chrome' --desc 'local'", flush=True)
        return False

    print(f"  Found browser: {browser_id}", flush=True)

    # Connect
    cmd = [
        "browser-act", "--session", SESSION,
        "browser", "open",
        browser_id,
        "https://maps.google.com"
    ]
    print(f"  Running: {' '.join(cmd)}", flush=True)
    r = subprocess.run(cmd, capture_output=True, text=True, timeout=20)

    if r.returncode == 0:
        print("✅ Browser session started", flush=True)
        return True
    else:
        print(f"❌ Connection failed: {r.stderr[:200]}", flush=True)
        return False

def bc(*args, t=30):
    """Call browser-act"""
    try:
        r = subprocess.run(
            ["browser-act", "--session", SESSION] + list(args),
            capture_output=True, text=True, timeout=t
        )
        return r.stdout, r.stderr, r.returncode
    except subprocess.TimeoutExpired:
        return "", "TIMEOUT", -1
    except Exception as e:
        return "", str(e), -1

AREAS = ["Kemang", "Pondok Indah", "Senopati", "Fatmawati",
         "Kuningan", "Cilandak", "Blok M", "Lebak Bulus"]
KEYWORDS = ["Kuliner"]

all_results = {}
parsed_entries = []

def search(query):
    """Search Google Maps dengan query"""
    print(f"\n▶ {query}", flush=True)

    # Navigate
    bc("navigate", "https://maps.google.com", t=12)
    time.sleep(1)

    # Type query
    bc("input", "3", query, t=8)
    time.sleep(0.3)
    bc("keys", "Enter", t=8)
    time.sleep(3)

    # Wait for stable
    bc("wait", "stable", "--timeout", "5000", t=12)

    # Scroll dengan JS (simple, no async)
    js = """
var p = document.querySelectorAll('.m6QErb.DxyBCb');
var target = null;
for (var i = 0; i < p.length; i++) {
  if (p[i].scrollHeight > 500) { target = p[i]; break; }
}
if (!target) {
  'no-panel';
} else {
  for (let s = 0; s < 10; s++) {
    target.scrollTop += 2500;
  }
  'scrolled';
}
"""
    bc("eval", js, t=15)
    time.sleep(1)

    # Extract markdown
    md, _, _ = bc("get", "markdown", t=10)
    ents = parse(md)
    print(f"  {len(ents)} hasil", flush=True)
    return ents

def parse(md):
    """Parse markdown hasil"""
    ents = []
    lines = [l.strip() for l in md.split('\n') if l.strip()]
    i = 0
    while i < len(lines):
        if i + 1 < len(lines) and re.match(r'^\d+[.,]\d+\(', lines[i+1]):
            e = extract(lines, i)
            if e: ents.append(e)
        i += 1
    return ents

def extract(lines, start):
    """Extract entry detail"""
    name = lines[start]
    rl = lines[start+1]
    rm = re.match(r'^(\d+[.,]\d+)', rl)
    rvm = re.search(r'\((\d+)\)', rl)
    e = {
        'name': name,
        'rating': float(rm.group(1).replace(',', '.')) if rm else 0,
        'reviews': int(rvm.group(1)) if rvm else 0,
        'address': '', 'phone': '', 'website': '', 'hours': '', 'category': ''
    }

    idx = start + 2
    while idx < len(lines) and idx < start + 20:
        line = lines[idx]

        # Stop if next entry
        if line.startswith(('PT', 'Restoran', 'Warung')) and idx + 1 < len(lines):
            if re.match(r'^\d+[.,]\d+\(', lines[idx + 1]):
                break

        # Website
        if 'Situs Web' in line:
            web = re.search(r'\]\((https?://[^\)]+)\)', line)
            if web: e['website'] = web.group(1)

        # Address
        if any(x in line for x in ['Kantor', 'Jalan', 'Jl.', 'Alamat']):
            addr = re.sub(r'[·☹\[\]\(\)]', '', line).strip()
            if addr and len(addr) > 5 and not addr.startswith('http'):
                e['address'] = addr[:80]

        # Hours & phone
        if ('Buka' in line or 'Tutup' in line) and '·' in line:
            parts = [p.strip() for p in line.split('·') if p.strip()]
            hp, pp = [], []
            for p in parts:
                if any(x in p for x in ['Buka', 'Tutup', 'pukul'] + [f'{h}.' for h in range(24)]):
                    hp.append(p)
                elif re.search(r'\(\d+\)', p) or re.match(r'^0\d', p) or '021' in p or '082' in p:
                    pp.append(p)
            e['hours'] = ' · '.join(hp) if hp else ''
            e['phone'] = ' · '.join(pp) if pp else ''

        idx += 1

    return e if e.get('name') else None

def dedup():
    """Fuzzy dedup"""
    print("\n🔄 Dedup...", end=" ", flush=True)
    unique = {}
    for results in all_results.values():
        for entry in results:
            nk = entry['name'].lower().strip()
            found = False
            for ek, ee in list(unique.items()):
                if fuzz.ratio(nk, ek) > 85:
                    merged = ee.copy()
                    for k in ['rating', 'reviews']:
                        if entry[k] > merged[k]: merged[k] = entry[k]
                    for k in ['address', 'phone', 'website', 'hours']:
                        if not merged[k] and entry.get(k): merged[k] = entry[k]
                    unique[ek] = merged
                    found = True
                    break
            if not found: unique[nk] = entry

    global parsed_entries
    parsed_entries = sorted(unique.values(), key=lambda x: (-x['rating'], -x['reviews']))
    total = sum(len(v) for v in all_results.values())
    print(f"{len(parsed_entries)} unique dari {total} raw", flush=True)

def xlsx(path):
    """Export ke XLSX"""
    print("📊 XLSX...", end=" ", flush=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hfont = Font(bold=True, color='FFFFFF', size=11)
    halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bdr = Border(left=Side(style='thin'), right=Side(style='thin'),
                 top=Side(style='thin'), bottom=Side(style='thin'))
    headers = ["No", "Nama", "Rating", "Review", "Alamat", "Telepon", "Jam", "Website"]
    cw = [5, 40, 8, 8, 40, 22, 22, 35]

    def fill(ws, data, hc):
        hf = PatternFill(start_color=hc, end_color=hc, fill_type="solid")
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hfont; c.fill = hf; c.alignment = halign; c.border = bdr
        for ri, e in enumerate(data, 2):
            ws.append([ri-1, e['name'], e['rating'], e['reviews'],
                      e['address'][:50], e['phone'][:30],
                      e['hours'][:25], e['website'] if e['website'] else '-'])
            for cell in ws[ri]: cell.border = bdr
            if not e['website']:
                ws.cell(row=ri, column=8).fill = \
                    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for ci, w in enumerate(cw, 1):
            ws.column_dimensions[chr(64+ci)].width = w
        ws.freeze_panes = "A2"

    fill(wb.create_sheet("Semua", 0), parsed_entries, "1F4E78")
    fill(wb.create_sheet("Ada Website", 1),
         [e for e in parsed_entries if e['website']], "2E7D32")
    fill(wb.create_sheet("Tanpa Website", 2),
         [e for e in parsed_entries if not e['website']], "C62828")
    wb.save(path)
    print(f"✅ {path}", flush=True)

def run():
    """Main"""
    print(f"\n{'='*60}", flush=True)
    print(f"🍜 KULINER SCOUT — Jakarta Selatan", flush=True)
    print(f"   Area: {len(AREAS)}", flush=True)
    print(f"{'='*60}\n", flush=True)

    # Check Chrome dulu
    if not check_chrome_ready():
        print("\n❌ Chrome gak running di localhost:9222", flush=True)
        print("   Start Chrome dulu: google-chrome --remote-debugging-port=9222 &", flush=True)
        return None

    # Init browser
    if not init_browser():
        print("❌ Gagal connect ke browser", flush=True)
        return None

    time.sleep(2)

    # Search
    for ai, area in enumerate(AREAS, 1):
        for ki, kw in enumerate(KEYWORDS, 1):
            query = f"{kw} {area}"
            try:
                results = search(query)
                all_results[query] = results
            except Exception as e:
                print(f"  ⚠️ Error: {str(e)[:60]}", flush=True)
            time.sleep(1)

    # Dedup & export
    if not all_results:
        print("\n❌ No results found", flush=True)
        return None

    dedup()
    path = "/home/ngome/Desktop/Kuliner_Jaksel.xlsx"
    xlsx(path)

    ww = sum(1 for e in parsed_entries if e['website'])
    print(f"\n{'='*60}", flush=True)
    print(f"✅ {len(parsed_entries)} unik | {ww} website | {len(parsed_entries)-ww} tanpa", flush=True)
    print(f"{'='*60}", flush=True)

    return path

if __name__ == "__main__":
    out = run()
    if out:
        print(f"\n✨ Output: {out}", flush=True)
    else:
        print(f"\n❌ Failed", flush=True)
        sys.exit(1)
