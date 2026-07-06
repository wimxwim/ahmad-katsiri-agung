#!/usr/bin/env python3
"""
Test version with mock data - no browser-act required
Output path: /home/ngome/Desktop/
"""

import sys
import re
import json
import time
from fuzzywuzzy import fuzz
from datetime import datetime
import openpyxl
from collections import defaultdict

class BusinessScraperMultiQuery:
    def __init__(self, location, sub_areas=None):
        self.location = location
        self.sub_areas = sub_areas or self._get_default_sub_areas(location)
        self.all_results = defaultdict(dict)
        self.parsed_entries = []

    def _get_default_sub_areas(self, location):
        """Default sub-areas untuk berbagai kota"""
        if "Jakarta Selatan" in location:
            return [
                "Pesanggrahan", "Mampang Prapatan", "Tebet",
                "Pasar Minggu", "Cipete", "Gandaria",
                "Kebayoran Baru", "Jagakarsa", "Senayan"
            ]
        return []

    def generate_mock_data(self):
        """Generate realistic mock data untuk testing"""
        print("🔍 Generating mock data for testing...")

        mock_entries = [
            {'name': 'PT Teknologi Indonesia', 'rating': 4.5, 'reviews': 43, 'address': 'Jl. Senayan No. 123, Jakarta Selatan', 'phone': '021-1234567', 'hours': 'Buka · 09.00', 'website': 'https://teknik-indonesia.com', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Maju Jaya Pratama', 'rating': 4.2, 'reviews': 28, 'address': 'Jl. Gandaria Utama No. 45, Jakarta Selatan', 'phone': '021-9876543', 'hours': 'Buka · 08.00', 'website': '', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Digital Solutions', 'rating': 4.8, 'reviews': 156, 'address': 'Jl. Kebayoran Baru No. 67, Jakarta Selatan', 'phone': '021-5555555', 'hours': 'Buka · 10.00', 'website': 'https://digital-solutions.id', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Konsultan Bisnis', 'rating': 3.9, 'reviews': 12, 'address': 'Jl. Tebet No. 89, Jakarta Selatan', 'phone': '021-2222222', 'hours': 'Buka · 09.30', 'website': 'https://konsultan-bisnis.co.id', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Transportasi Cepat', 'rating': 4.1, 'reviews': 67, 'address': 'Jl. Cipete No. 101, Jakarta Selatan', 'phone': '021-7777777', 'hours': 'Buka · 06.00', 'website': '', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Media Kreatif', 'rating': 4.6, 'reviews': 89, 'address': 'Jl. Mampang Prapatan No. 234, Jakarta Selatan', 'phone': '021-8888888', 'hours': 'Buka · 09.00', 'website': 'https://media-kreatif.com', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Bangun Konstruksi', 'rating': 4.3, 'reviews': 45, 'address': 'Jl. Pasar Minggu No. 345, Jakarta Selatan', 'phone': '021-4444444', 'hours': 'Buka · 08.00', 'website': '', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Energi Mandiri', 'rating': 4.7, 'reviews': 112, 'address': 'Jl. Jagakarsa No. 456, Jakarta Selatan', 'phone': '021-6666666', 'hours': 'Buka · 07.00', 'website': 'https://energi-mandiri.net', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Edukasi Global', 'rating': 4.4, 'reviews': 78, 'address': 'Jl. Pesanggrahan No. 567, Jakarta Selatan', 'phone': '021-3333333', 'hours': 'Buka · 09.00', 'website': 'https://edukasi-global.org', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Logistik Semesta', 'rating': 4.0, 'reviews': 34, 'address': 'Jl. Gandaria No. 678, Jakarta Selatan', 'phone': '021-9191919', 'hours': 'Buka · 06.00', 'website': '', 'category': 'Kantor Perusahaan'},
        ]

        all_mock = []
        for i, entry in enumerate(mock_entries):
            all_mock.append(entry)

        for i in range(15):
            for entry in mock_entries[:5]:
                variant = entry.copy()
                variant['name'] = variant['name'] + f" - Cabang {i+1}" if i % 3 != 0 else variant['name']
                variant['reviews'] = variant['reviews'] + (i % 10)
                all_mock.append(variant)

        for i in range(20):
            entry = mock_entries[i % len(mock_entries)].copy()
            entry['website'] = ''
            entry['name'] = f"PT Lokal {i+1}"
            all_mock.append(entry)

        self.all_results['PT Jakarta Selatan'] = all_mock[:len(mock_entries) + 15*5 + 20]
        print(f"  ✅ Generated {len(self.all_results['PT Jakarta Selatan'])} mock entries")

    def deduplicate_entries(self):
        """Fuzzy deduplication across all results"""
        print("\n🔄 Deduplicating entries...")

        unique_entries = {}
        for results in self.all_results.values():
            for entry in results:
                key = entry['name'].lower().strip()
                is_dup = False
                for existing_key in unique_entries:
                    if fuzz.ratio(key, existing_key) > 85 or fuzz.partial_ratio(key, existing_key) > 90:
                        is_dup = True
                        break
                if not is_dup:
                    unique_entries[key] = entry

        self.parsed_entries = list(unique_entries.values())
        print(f"  ✅ {len(self.parsed_entries)} unique entries found (from {sum(len(v) for v in self.all_results.values())} total)")

    def generate_xlsx_report(self, output_path):
        """Generate XLSX with 3 sheets"""
        print("\n📊 Generating XLSX report...")

        wb = openpyxl.Workbook()
        hfont = Font(bold=True, color='FFFFFF', size=11)
        halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bdr = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ['No', 'Nama PT', 'Rating', 'Jumlah Review', 'Alamat', 'Telepon', 'Jam Operasional', 'Website', 'Status Website']
        cw = [5, 42, 10, 15, 55, 22, 30, 40, 18]

        def write_sheet(ws, data, hfill):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = bdr
            for ri, e in enumerate(data, 2):
                vals = [ri-1, e['name'], e['rating'], e['reviews'],
                        e['address'], e['phone'], e['hours'],
                        e['website'] if e.get('website') else '-',
                        'Ada Website' if e.get('website') else 'Tanpa Website']
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row=ri, column=ci, value=v)
                    c.border = bdr
                sc = ws.cell(row=ri, column=9)
                sc.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') if e.get('website') else PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            for ci, w in enumerate(cw, 1):
                ws.column_dimensions[chr(64+ci)].width = w

        fill_blue = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        fill_green = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
        fill_red = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')

        ws_all = wb.active; ws_all.title = "Semua PT"
        write_sheet(ws_all, self.parsed_entries, fill_blue)

        ws_w = wb.create_sheet("PT dengan Website")
        write_sheet(ws_w, [e for e in self.parsed_entries if e.get('website')], fill_green)

        ws_wo = wb.create_sheet("PT Tanpa Website")
        write_sheet(ws_wo, [e for e in self.parsed_entries if not e.get('website')], fill_red)

        wb.save(output_path)
        print(f"  ✅ XLSX saved to: {output_path}")

    def run_complete_workflow(self, output_path):
        """Run the complete workflow: generate -> dedup -> export"""
        self.generate_mock_data()
        self.deduplicate_entries()
        self.generate_xlsx_report(output_path)
        return {
            'total_raw': sum(len(v) for v in self.all_results.values()),
            'unique': len(self.parsed_entries),
            'with_website': len([e for e in self.parsed_entries if e.get('website')]),
            'without_website': len([e for e in self.parsed_entries if not e.get('website')]),
            'output_path': output_path
        }

if __name__ == "__main__":
    location = "Jakarta Selatan"
    print("=" * 80)
    print(f"MULTI-QUERY LOCATION BUSINESS SCRAPER (TEST MODE)")
    print(f"Location: {location}")
    print(f"Sub-areas: {len(BusinessScraperMultiQuery(location)._get_default_sub_areas(location))}")
    print("=" * 80)

    scraper = BusinessScraperMultiQuery(location)
    output = scraper.run_complete_workflow("/home/ngome/Desktop/Test_Multi_Query_Output.xlsx")

    print("\n" + "=" * 80)
    print("TEST RESULT SUMMARY")
    print("=" * 80)
    print(f"  Total raw entries : {output['total_raw']}")
    print(f"  After dedup       : {output['unique']}")
    print(f"  With website      : {output['with_website']}")
    print(f"  Without website   : {output['without_website']}")
    print(f"  Dedup rate        : {(1 - output['unique']/output['total_raw'])*100:.1f}%")
    print(f"  Output            : {output['output_path']}")
    print("=" * 80)
