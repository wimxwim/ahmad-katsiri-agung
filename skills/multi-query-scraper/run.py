#!/usr/bin/env python3
"""
Multi-Query Location Business Scraper
Hermes Skill Version
"""

import re
import json
import time
import subprocess
from fuzzywuzzy import fuzz
from datetime import datetime
import openpyxl
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class BusinessScraperMultiQuery:
    def __init__(self, location, sub_areas=None):
        self.location = location
        self.sub_areas = sub_areas or self._get_default_sub_areas(location)
        self.all_results = defaultdict(dict)
        self.parsed_entries = []
        
    def _get_default_sub_areas(self, location):
        """Default sub-areas untuk Jakarta Selatan"""
        if "Jakarta Selatan" in location:
            return [
                "Pesanggrahan", "Mampang Prapatan", "Tebet", 
                "Pasar Minggu", "Cipete", "Gandaria", 
                "Kebayoran Baru", "Jagakarsa"
            ]
        return []
    
    def search_all_sub_areas(self):
        """Execute multi-query search"""
        print(f"🔍 Searching {len(self.sub_areas)} sub-areas in {self.location}")
        
        for idx, sub_area in enumerate(self.sub_areas, 1):
            print(f"\n[{idx}/{len(self.sub_areas)}] Searching: {sub_area}")
            
            queries = [
                f"PT {sub_area}",
                f"Perusahaan {sub_area}",
            ]
            
            for query in queries:
                try:
                    results = self._search_maps(query)
                    self.all_results[query] = results
                    print(f"  ✅ Query '{query}': {len(results)} results")
                except Exception as e:
                    print(f"  ⚠️ Query '{query}' failed: {e}")
            
            time.sleep(1)
    
    def _search_maps(self, query):
        """Search Google Maps dengan aggressive scroll"""
        try:
            # Navigate
            subprocess.run(["browser-act", "--session", "scraper", "navigate", "https://maps.google.com"], 
                         capture_output=True, timeout=10)
            time.sleep(1)
            
            # Search
            subprocess.run(["browser-act", "--session", "scraper", "click", "3"], 
                         capture_output=True, timeout=5)
            subprocess.run(["browser-act", "--session", "scraper", "input", "3", query], 
                         capture_output=True, timeout=5)
            subprocess.run(["browser-act", "--session", "scraper", "keys", "Enter"], 
                         capture_output=True, timeout=5)
            
            # Wait
            subprocess.run(["browser-act", "--session", "scraper", "wait", "stable", "--timeout", "5000"], 
                         capture_output=True, timeout=10)
            
            # Smart scroll dengan JS
            scroll_js = """
            var panels = document.querySelectorAll('.m6QErb.DxyBCb');
            for (var i = 0; i < panels.length; i++) {
              panels[i].scrollTop = 0;
            }
            for (let step = 0; step < 15; step++) {
              for (var p = 0; p < panels.length; p++) {
                if (panels[p].scrollHeight > 2000) {
                  panels[p].scrollTop += 2000;
                }
              }
              await new Promise(r => setTimeout(r, 300));
            }
            'scrolled'
            """
            
            subprocess.run(["browser-act", "--session", "scraper", "eval", scroll_js], 
                         capture_output=True, timeout=30)
            
            # Extract
            result = subprocess.run(["browser-act", "--session", "scraper", "get", "markdown"], 
                                  capture_output=True, text=True, timeout=10)
            
            return self._parse_results(result.stdout)
        except Exception as e:
            print(f"    Search error: {e}")
            return []
    
    def _parse_results(self, markdown):
        """Parse markdown results"""
        entries = []
        lines = [l.strip() for l in markdown.split('\n') if l.strip()]
        
        i = 0
        while i < len(lines):
            line = lines[i]
            
            if line.startswith('PT') and i + 1 < len(lines):
                next_line = lines[i + 1]
                
                if re.match(r'^\d+[.,]\d+\(\d+\)', next_line):
                    entry = self._extract_entry_details(lines, i)
                    if entry:
                        entries.append(entry)
                    i += 1
                else:
                    i += 1
            else:
                i += 1
        
        return entries
    
    def _extract_entry_details(self, lines, start_idx):
        """Extract detail dari entry"""
        name = lines[start_idx]
        rating_line = lines[start_idx + 1]
        
        rating_match = re.match(r'^(\d+[.,]\d+)', rating_line)
        review_match = re.search(r'\((\d+)\)', rating_line)
        
        entry = {
            'name': name,
            'rating': float(rating_match.group(1).replace(',', '.')) if rating_match else 0,
            'reviews': int(review_match.group(1)) if review_match else 0,
            'address': '',
            'phone': '',
            'hours': '',
            'website': '',
        }
        
        idx = start_idx + 2
        while idx < len(lines) and idx < start_idx + 20:
            line = lines[idx]
            
            if line.startswith('PT'):
                break
            
            if 'Situs Web' in line:
                web_match = re.search(r'\]\((https?://[^\)]+)\)', line)
                if web_match:
                    entry['website'] = web_match.group(1)
            
            if 'Kantor Perusahaan' in line:
                addr = line.replace('Kantor Perusahaan · ', '').replace('Kantor Perusahaan', '')
                addr = re.sub(r'[·☹]', '', addr).strip()
                if addr and len(addr) > 5:
                    entry['address'] = addr
            
            if ('Buka' in line or 'Tutup' in line) and '·' in line:
                parts = [p.strip() for p in line.split('·') if p.strip()]
                hours_parts = []
                phone_parts = []
                
                for p in parts:
                    if any(x in p for x in ['Buka', 'Tutup', 'pukul', '09.', '17.', '16.', '15.']):
                        hours_parts.append(p)
                    elif re.search(r'\(\d+\)', p) or re.match(r'^0\d', p) or '021' in p or '082' in p:
                        phone_parts.append(p)
                
                entry['hours'] = ' · '.join(hours_parts) if hours_parts else ''
                entry['phone'] = ' · '.join(phone_parts) if phone_parts else ''
            
            idx += 1
        
        return entry if entry.get('name') else None
    
    def deduplicate_entries(self):
        """Fuzzy deduplication"""
        print("\n🔄 Deduplicating entries...")
        
        unique_entries = {}
        
        for results in self.all_results.values():
            for entry in results:
                name_key = entry['name'].lower().strip()
                
                found_match = False
                for existing_name, existing_entry in unique_entries.items():
                    if fuzz.ratio(name_key, existing_name) > 90:
                        merged = self._merge_entries(existing_entry, entry)
                        unique_entries[existing_name] = merged
                        found_match = True
                        break
                
                if not found_match:
                    unique_entries[name_key] = entry
        
        self.parsed_entries = list(unique_entries.values())
        self.parsed_entries.sort(key=lambda x: (-x['rating'], -x['reviews']))
        
        print(f"  ✅ {len(self.parsed_entries)} unique entries found")
    
    def _merge_entries(self, entry1, entry2):
        """Merge entries dengan smart logic"""
        merged = entry1.copy()
        
        if entry2['rating'] > merged['rating']:
            merged['rating'] = entry2['rating']
        
        if entry2['reviews'] > merged['reviews']:
            merged['reviews'] = entry2['reviews']
        
        for key in ['address', 'phone', 'website', 'hours']:
            if not merged[key] and entry2.get(key):
                merged[key] = entry2[key]
        
        return merged
    
    def generate_xlsx_report(self, output_path):
        """Generate XLSX report"""
        print(f"\n📊 Generating XLSX report...")
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Sheet 1: All entries
        ws_all = wb.create_sheet("Semua PT", 0)
        self._populate_sheet(ws_all, self.parsed_entries)
        
        # Sheet 2: With website
        with_web = [e for e in self.parsed_entries if e['website']]
        ws_web = wb.create_sheet("PT Dengan Website", 1)
        self._populate_sheet(ws_web, with_web)
        
        # Sheet 3: Without website
        without_web = [e for e in self.parsed_entries if not e['website']]
        ws_prospect = wb.create_sheet("PT Tanpa Website (Prospek)", 2)
        self._populate_sheet(ws_prospect, without_web, color_red=True)
        
        wb.save(output_path)
        print(f"  ✅ Saved: {output_path}")
    
    def _populate_sheet(self, ws, entries, color_red=False):
        """Populate worksheet"""
        headers = ["No", "Nama PT", "Rating", "Review", "Alamat", "Telepon", "Jam", "Website"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for idx, entry in enumerate(entries, 2):
            ws.append([
                idx - 1,
                entry['name'],
                entry['rating'],
                entry['reviews'],
                entry['address'][:50] if entry['address'] else '',
                entry['phone'][:30] if entry['phone'] else '',
                entry['hours'][:25] if entry['hours'] else '',
                entry['website'] if entry['website'] else '-'
            ])
            
            if color_red:
                ws[f'H{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 30
        
        ws.freeze_panes = "A2"
    
    def run_complete_workflow(self, output_path):
        """Run complete workflow"""
        print(f"{'='*80}")
        print(f"🚀 MULTI-QUERY LOCATION BUSINESS SCRAPER")
        print(f"Location: {self.location}")
        print(f"Sub-areas: {len(self.sub_areas)}")
        print(f"{'='*80}\n")
        
        self.search_all_sub_areas()
        self.deduplicate_entries()
        self.generate_xlsx_report(output_path)
        
        total_results = sum(len(v) for v in self.all_results.values())
        with_website = sum(1 for e in self.parsed_entries if e['website'])
        
        print(f"\n{'='*80}")
        print(f"📈 STATISTICS")
        print(f"  Total raw results: {total_results}")
        print(f"  Unique entries: {len(self.parsed_entries)}")
        print(f"  With website: {with_website} ({with_website*100//len(self.parsed_entries) if self.parsed_entries else 0}%)")
        print(f"  Without website: {len(self.parsed_entries) - with_website}")
        print(f"{'='*80}\n")
        
        return output_path

if __name__ == "__main__":
    scraper = BusinessScraperMultiQuery("Jakarta Selatan")
    output = scraper.run_complete_workflow("/home/ngome/Desktop/Hermes_Multi_Query_Results.xlsx")
    print(f"✅ Complete! Output: {output}")
