#!/usr/bin/env python3
"""
Multi-Query Location Business Scraper — Fixed Version
"""
import re, json, time, subprocess, sys
from fuzzywuzzy import fuzz
from datetime import datetime
import openpyxl
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

S = lambda *a, **kw: subprocess.run(["browser-act", "--session", "scraper"] + list(a),
    capture_output=True, text=True, timeout=kw.get("t", 15))

class BusinessScraperMultiQuery:
    def __init__(self, location, sub_areas=None):
        self.location = location
        self.sub_areas = sub_areas or self._get_default_sub_areas(location)
        self.all_results = defaultdict(dict)
        self.parsed_entries = []

    def _get_default_sub_areas(self, location):
        if "Jakarta Selatan" in location:
            return ["Pesanggrahan", "Mampang Prapatan", "Tebet", "Pasar Minggu",
                    "Cipete", "Gandaria", "Kebayoran Baru", "Jagakarsa"]
        return []

    def _ensure_session(self):
        # Close any stale session, open fresh browser
        S("session", "close", "scraper", t=3)
        r = S("browser", "open", "chrome_local_98045233188569264",
              "https://maps.google.com", t=15)
        if r.returncode != 0:
            print(f"  ⚠️ Browser open failed: {r.stderr.strip()[:100]}")
            return False
        time.sleep(2)
        return True

    def search_all_sub_areas(self):
        print(f"🔍 Searching {len(self.sub_areas)} sub-areas in {self.location}\n")
        if not self._ensure_session():
            return

        for idx, sub_area in enumerate(self.sub_areas, 1):
            queries = [f"PT {sub_area}", f"Perusahaan {sub_area}"]
            for qidx, query in enumerate(queries, 1):
                print(f"[{idx}/{len(self.sub_areas)}][{qidx}/{len(queries)}] {query}")
                try:
                    results = self._search_maps(query)
                    self.all_results[query] = results
                    print(f"  → {len(results)} results")
                except Exception as e:
                    print(f"  ⚠️ Error: {e}")
                time.sleep(1)

    def _bc(self, *args, t=15):
        r = S(*args, t=t)
        if r.returncode != 0:
            raise RuntimeError(f"browser-act {' '.join(args)}: {r.stderr.strip()[:100]}")
        return r.stdout

    def _search_maps(self, query):
        self._bc("navigate", "https://maps.google.com", t=10)
        time.sleep(1.5)
        self._bc("input", "3", query, t=8)
        time.sleep(0.5)
        self._bc("keys", "Enter", t=8)
        time.sleep(3)
        self._bc("wait", "stable", "--timeout", "5000", t=12)

        # Smart scroll — async IIFE
        scroll_js = """
(async () => {
  var panels = document.querySelectorAll('.m6QErb.DxyBCb');
  var target = null;
  for (var i = 0; i < panels.length; i++) {
    if (panels[i].scrollHeight > 2000) {
      target = panels[i];
      break;
    }
  }
  if (!target) return 'no-panel';
  for (let step = 0; step < 12; step++) {
    target.scrollTop += 3000;
    await new Promise(r => setTimeout(r, 400));
  }
  return 'scrolled-' + target.scrollTop;
})()
"""
        r = S("eval", scroll_js, t=45)
        self._bc("wait", "stable", "--timeout", "3000", t=10)

        r2 = S("get", "markdown", t=10)
        return self._parse_results(r2.stdout)

    def _parse_results(self, markdown):
        entries = []
        lines = [l.strip() for l in markdown.split('\n') if l.strip()]
        i = 0
        while i < len(lines):
            line = lines[i]
            if line.startswith('PT') and i + 1 < len(lines):
                nl = lines[i + 1]
                if re.match(r'^\d+[.,]\d+\(', nl):
                    entry = self._extract_entry_details(lines, i)
                    if entry: entries.append(entry)
            i += 1
        return entries

    def _extract_entry_details(self, lines, start_idx):
        name = lines[start_idx]
        rl = lines[start_idx + 1]
        rm = re.match(r'^(\d+[.,]\d+)', rl)
        rvm = re.search(r'\((\d+)\)', rl)
        entry = {'name': name, 'rating': float(rm.group(1).replace(',', '.')) if rm else 0,
                 'reviews': int(rvm.group(1)) if rvm else 0,
                 'address': '', 'phone': '', 'hours': '', 'website': ''}
        idx = start_idx + 2
        while idx < len(lines) and idx < start_idx + 20:
            line = lines[idx]
            if line.startswith('PT'): break
            if 'Situs Web' in line:
                wm = re.search(r'\]\((https?://[^\)]+)\)', line)
                if wm: entry['website'] = wm.group(1)
            if 'Kantor Perusahaan' in line:
                a = line.replace('Kantor Perusahaan · ', '').replace('Kantor Perusahaan', '')
                a = re.sub(r'[·☹]', '', a).strip()
                if a and len(a) > 5: entry['address'] = a
            if ('Buka' in line or 'Tutup' in line) and '·' in line:
                parts = [p.strip() for p in line.split('·') if p.strip()]
                hp, pp = [], []
                for p in parts:
                    if any(x in p for x in ['Buka', 'Tutup', 'pukul', '09.', '17.', '16.', '15.']): hp.append(p)
                    elif re.search(r'\(\d+\)', p) or re.match(r'^0\d', p) or '021' in p or '082' in p: pp.append(p)
                entry['hours'] = ' · '.join(hp) if hp else ''
                entry['phone'] = ' · '.join(pp) if pp else ''
            idx += 1
        return entry if entry.get('name') else None

    def deduplicate_entries(self):
        print("\n🔄 Deduplicating...")
        unique = {}
        for results in self.all_results.values():
            for entry in results:
                nk = entry['name'].lower().strip()
                found = False
                for ek, ee in list(unique.items()):
                    if fuzz.ratio(nk, ek) > 90:
                        merged = ee.copy()
                        if entry['rating'] > merged['rating']: merged['rating'] = entry['rating']
                        if entry['reviews'] > merged['reviews']: merged['reviews'] = entry['reviews']
                        for k in ['address', 'phone', 'website', 'hours']:
                            if not merged[k] and entry.get(k): merged[k] = entry[k]
                        unique[ek] = merged
                        found = True
                        break
                if not found: unique[nk] = entry
        self.parsed_entries = sorted(unique.values(), key=lambda x: (-x['rating'], -x['reviews']))
        total = sum(len(v) for v in self.all_results.values())
        print(f"  {len(self.parsed_entries)} unique from {total} raw")

    def generate_xlsx_report(self, output_path):
        print("\n📊 XLSX...")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        hfont = Font(bold=True, color='FFFFFF', size=11)
        halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bdr = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        headers = ["No", "Nama PT", "Rating", "Review", "Alamat", "Telepon", "Jam", "Website"]
        cw = [5, 38, 8, 8, 40, 22, 22, 35]

        def fill(ws, data, hl, hc):
            hfill = PatternFill(start_color=hl, end_color=hl, fill_type="solid")
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = bdr
            for ri, e in enumerate(data, 2):
                ws.append([ri-1, e['name'], e['rating'], e['reviews'],
                          e['address'][:50], e['phone'][:30], e['hours'][:25],
                          e['website'] if e['website'] else '-'])
                for cell in ws[ri]: cell.border = bdr
                if not e['website'] and hc == "red":
                    ws.cell(row=ri, column=8).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for ci, w in enumerate(cw, 1):
                ws.column_dimensions[chr(64+ci)].width = w
            ws.freeze_panes = "A2"

        fill(wb.create_sheet("Semua PT", 0), self.parsed_entries, "1F4E78", "blue")
        fill(wb.create_sheet("PT Dengan Website", 1),
             [e for e in self.parsed_entries if e['website']], "2E7D32", "green")
        fill(wb.create_sheet("PT Tanpa Website", 2),
             [e for e in self.parsed_entries if not e['website']], "C62828", "red")
        wb.save(output_path)
        print(f"  ✅ {output_path}")

    def run(self, output_path):
        print(f"{'='*70}")
        print(f"🚀 MULTI-QUERY LOCATION SCRAPER")
        print(f"Location: {self.location} | Sub-areas: {len(self.sub_areas)}")
        print(f"{'='*70}\n")
        self.search_all_sub_areas()
        self.deduplicate_entries()
        self.generate_xlsx_report(output_path)
        ww = sum(1 for e in self.parsed_entries if e['website'])
        print(f"\n{'='*70}")
        print(f"📈 STATS: {len(self.parsed_entries)} unique | {ww} ada website | {len(self.parsed_entries)-ww} tanpa website")
        print(f"{'='*70}")

if __name__ == "__main__":
    s = BusinessScraperMultiQuery("Jakarta Selatan")
    s.run("/home/ngome/Desktop/Hermes_Multi_Query_Results.xlsx")
