#!/usr/bin/env python3
"""
Test version with mock data - no browser-act required
"""

import sys
sys.path.insert(0, '/home/ngome/.agents/skills/browser-act')

import re
import json
import time
from fuzzywuzzy import fuzz
from datetime import datetime
import openpyxl
from collections import defaultdict

class BusinessScraperMultiQuery:
    def __init__(self, location, sub_areas=None):
        self.location = location
        self.sub_areas = sub_areas or self._get_default_sub_areas(location)
        self.all_results = defaultdict(dict)
        self.parsed_entries = []

    def _get_default_sub_areas(self, location):
        """Default sub-areas untuk berbagai kota"""
        if "Jakarta Selatan" in location:
            return [
                "Pesanggrahan", "Mampang Prapatan", "Tebet",
                "Pasar Minggu", "Cipete", "Gandaria",
                "Kebayoran Baru", "Jagakarsa", "Senayan"
            ]
        return []

    def generate_mock_data(self):
        """Generate realistic mock data untuk testing"""
        print("🔍 Generating mock data for testing...")

        mock_entries = [
            {'name': 'PT Teknologi Indonesia', 'rating': 4.5, 'reviews': 43, 'address': 'Jl. Senayan No. 123, Jakarta Selatan', 'phone': '021-1234567', 'hours': 'Buka · 09.00', 'website': 'https://teknik-indonesia.com', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Maju Jaya Pratama', 'rating': 4.2, 'reviews': 28, 'address': 'Jl. Gandaria Utama No. 45, Jakarta Selatan', 'phone': '021-9876543', 'hours': 'Buka · 08.00', 'website': '', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Digital Solutions', 'rating': 4.8, 'reviews': 156, 'address': 'Jl. Kebayoran Baru No. 67, Jakarta Selatan', 'phone': '021-5555555', 'hours': 'Buka · 10.00', 'website': 'https://digital-solutions.id', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Konsultan Bisnis', 'rating': 3.9, 'reviews': 12, 'address': 'Jl. Tebet No. 89, Jakarta Selatan', 'phone': '021-2222222', 'hours': 'Buka · 09.30', 'website': 'https://konsultan-bisnis.co.id', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Transportasi Cepat', 'rating': 4.1, 'reviews': 67, 'address': 'Jl. Cipete No. 101, Jakarta Selatan', 'phone': '021-7777777', 'hours': 'Buka · 06.00', 'website': '', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Media Kreatif', 'rating': 4.6, 'reviews': 89, 'address': 'Jl. Mampang Prapatan No. 234, Jakarta Selatan', 'phone': '021-8888888', 'hours': 'Buka · 09.00', 'website': 'https://media-kreatif.com', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Bangun Konstruksi', 'rating': 4.3, 'reviews': 45, 'address': 'Jl. Pasar Minggu No. 345, Jakarta Selatan', 'phone': '021-4444444', 'hours': 'Buka · 08.00', 'website': '', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Energi Mandiri', 'rating': 4.7, 'reviews': 112, 'address': 'Jl. Jagakarsa No. 456, Jakarta Selatan', 'phone': '021-6666666', 'hours': 'Buka · 07.00', 'website': 'https://energi-mandiri.net', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Edukasi Global', 'rating': 4.4, 'reviews': 78, 'address': 'Jl. Pesanggrahan No. 567, Jakarta Selatan', 'phone': '021-3333333', 'hours': 'Buka · 09.00', 'website': 'https://edukasi-global.org', 'category': 'Kantor Perusahaan'},
            {'name': 'PT Logistik Semesta', 'rating': 4.0, 'reviews': 34, 'address': 'Jl. Gandaria No. 678, Jakarta Selatan', 'phone': '021-9191919', 'hours': 'Buka · 06.00', 'website': '', 'category': 'Kantor Perusahaan'},
        ]

        # Generate 150+ entries with variations
        all_mock = []
        for i, entry in enumerate(mock_entries):
            all_mock.append(entry)

        # Add more entries with duplicates to test dedup
        for i in range(15):
            for entry in mock_entries[:5]:
                variant = entry.copy()
                variant['name'] = variant['name'] + f" - Cabang {i+1}" if i % 3 != 0 else variant['name']
                variant['reviews'] = variant['reviews'] + (i % 10)
                all_mock.append(variant)

        # Add entries without website
        for i in range(20):
            entry = mock_entries[i % len(mock_entries)].copy()
            entry['website'] = ''
            entry['name'] = f"PT Lokal {i+1}"
            all_mock.append(entry)

        self.all_results['PT Jakarta Selatan'] = all_mock[:160]
        print(f"  ✅ Generated {len(all_mock[:160])} mock entries")

    def deduplicate_entries(self):
        """Fuzzy deduplication across all results"""
        print("\n🔄 Deduplicating entries...")

        unique_entries = {}

        for results in self.all_results.values():
            for entry in results:
                name_key = entry['name'].lower().strip()

                # Check fuzzy match
                found_match = False
                for existing_name, existing_entry in unique_entries.items():
                    similarity = fuzz.ratio(name_key, existing_name)

                    if similarity > 85:  # 85% similarity threshold
                        # Merge entries (keep most complete data)
                        merged = self._merge_entries(existing_entry, entry)
                        unique_entries[existing_name] = merged
                        found_match = True
                        break

                if not found_match:
                    unique_entries[name_key] = entry

        self.parsed_entries = list(unique_entries.values())
        self.parsed_entries.sort(key=lambda x: (-x['rating'], -x['reviews']))

        print(f"  ✅ {len(self.parsed_entries)} unique entries found")

    def _merge_entries(self, entry1, entry2):
        """Merge dua entry dengan smart logic"""
        merged = entry1.copy()

        # Keep highest rating
        if entry2['rating'] > merged['rating']:
            merged['rating'] = entry2['rating']

        # Keep highest review count
        if entry2['reviews'] > merged['reviews']:
            merged['reviews'] = entry2['reviews']

        # Fill missing data from entry2
        for key in ['address', 'phone', 'website', 'hours']:
            if not merged[key] and entry2.get(key):
                merged[key] = entry2[key]

        return merged

    def generate_xlsx_report(self, output_path):
        """Generate professional XLSX report"""
        print(f"\n📊 Generating XLSX report...")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Sheet 1: All entries
        ws_all = wb.create_sheet("Semua PT", 0)
        self._populate_sheet(ws_all, self.parsed_entries)

        # Sheet 2: With website
        with_web = [e for e in self.parsed_entries if e['website']]
        ws_web = wb.create_sheet("PT Dengan Website", 1)
        self._populate_sheet(ws_web, with_web)

        # Sheet 3: Without website (prospects)
        without_web = [e for e in self.parsed_entries if not e['website']]
        ws_prospect = wb.create_sheet("PT Tanpa Website (Prospek)", 2)
        self._populate_sheet(ws_prospect, without_web, color_red=True)

        wb.save(output_path)
        print(f"  ✅ Saved: {output_path}")

    def _populate_sheet(self, ws, entries, color_red=False):
        """Populate worksheet dengan styling"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        headers = ["No", "Nama PT", "Rating", "Review", "Alamat", "Telepon", "Jam", "Website"]
        ws.append(headers)

        # Header styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Data rows
        for idx, entry in enumerate(entries, 2):
            ws.append([
                idx - 1,
                entry['name'],
                entry['rating'],
                entry['reviews'],
                entry['address'][:50],
                entry['phone'][:30],
                entry['hours'][:25],
                entry['website'] or '-'
            ])

            # Row styling
            if color_red:
                ws[f'H{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 30

        ws.freeze_panes = "A2"

    def run_complete_workflow(self, output_path):
        """Run complete scraping workflow"""
        print(f"{'='*80}")
        print(f"MULTI-QUERY LOCATION BUSINESS SCRAPER (TEST MODE)")
        print(f"Location: {self.location}")
        print(f"Sub-areas: {len(self.sub_areas)}")
        print(f"{'='*80}\n")

        self.generate_mock_data()
        self.deduplicate_entries()
        self.generate_xlsx_report(output_path)

        # Statistics
        total_results = sum(len(v) for v in self.all_results.values())
        with_website = sum(1 for e in self.parsed_entries if e['website'])

        print(f"\n{'='*80}")
        print(f"STATISTICS")
        print(f"  Total raw results: {total_results}")
        print(f"  Unique entries: {len(self.parsed_entries)}")
        print(f"  With website: {with_website} ({with_website*100//len(self.parsed_entries) if self.parsed_entries else 0}%)")
        print(f"  Without website: {len(self.parsed_entries) - with_website}")
        print(f"{'='*80}\n")

        return output_path

# Main execution
if __name__ == "__main__":
    scraper = BusinessScraperMultiQuery("Jakarta Selatan")
    output = scraper.run_complete_workflow("/home/ngome/Desktop/Test_Multi_Query_Output.xlsx")
    print(f"Complete! Output: {output}")
